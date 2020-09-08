'''
Created by Ethan Guinn
Last update: 9/8/2020

This program takes files from the "Manually_Classified" folder and uses the classifications to estimate features for
one of 19 statement types. The program then takes CSVs of dialogues in the "Unclssified" folder and assigns tags to
each line in the dialogue statements.

Revised code to read and write sample texts.
'''

import nltk, chardet, os, re, xlrd, timeit
from openpyxl import load_workbook

start_time = timeit.default_timer()

# Identify the encoding of a given file
def find_encoding(filename):
    enc = chardet.detect(open(filename, 'rb').read())['encoding']
    return enc

# Dictionary holding the therapist sentences as keys and their tag as a value
therapist_sentences = {}

# Import all .xlsx files
os.chdir(r"Manually_Classified")
filenames = [file for file in os.listdir() if re.search(r"\.xlsx", file)]

print('\nProcessing manually classified files . . .')

for filename in filenames:
    # print(filename)
    # Identify the therapist's sentences
    all_sheets = xlrd.open_workbook(filename)
    sheet = all_sheets.sheet_by_index(0)
    is_therapist = False
    for row_idx in range(sheet.nrows):
        dcell = sheet.cell(row_idx, 0).value # Cell with clause
        tagcell = sheet.cell(row_idx, 6).value # Cell with tag
        # Therapist's dialogues are marked with a prefix 'T' in the related cells
        if len(dcell) > 14:
            if dcell[14] == 'T':
                therapist_sentences[dcell[17:]] = tagcell
                is_therapist = True
            elif dcell[14] in ['M', 'F']:
                is_therapist = False
            elif is_therapist:
                therapist_sentences[dcell] = tagcell

# Features of the dialogue are defined here.
def dialogue_act_features(sentence):
    features = {}
    for word in nltk.word_tokenize(sentence):
        features['contains({})'.format(word.lower())] = True
    return features

# Builds training and test sets to calculate the accuracy of the predictions.
print('\nBuilding feature sets . . .')
featuresets = [(dialogue_act_features(sentence), therapist_sentences[sentence]) for sentence in therapist_sentences]
size = int(len(featuresets) * 0.6)
print("Size: ", size)
train_set, test_set = featuresets[size:], featuresets[:size]
classifier = nltk.DecisionTreeClassifier.train(train_set)
print('Tested accuracy: ', round(nltk.classify.accuracy(classifier, test_set), 4))

# Uncomment to test a sentence of your own in the terminal
# print(classifier.classify(dialogue_act_features(input('\nEnter sentence here: '))))

# Make predictions and save to new .xlsx files
os.chdir(r"../Unclassified")
filenames = [file for file in os.listdir() if re.search(r"\.xlsx", file)]

print('\nPredicting tags for files in folder "Unclassified" . . .')

for filename in filenames:
    print(filename)
    # ftitle = str(filename)[:-5]
    # NOTE: Openpyxl indecies are one-based, not zero-based; therefore, the cell locations we put into the sheets are one unit more.
    workbook = load_workbook(filename)
    worksheet = workbook.active
    # Identify the therapist's sentences
    all_sheets = xlrd.open_workbook(filename)
    sheet = all_sheets.sheet_by_index(0)
    is_therapist = False
    for row_idx in range(sheet.nrows):
        dcell = sheet.cell(row_idx, 0).value
        # Therapist's dialogues are marked with a prefix 'T' in the related cells
        if len(dcell) > 11:
            if dcell[11] == 'T':
                tag = classifier.classify(dialogue_act_features(dcell[14:]))
                worksheet.cell(row_idx + 1, 7, tag)
                is_therapist = True
            elif dcell[11] in ['M', 'F']:
                is_therapist = False
            elif is_therapist:
                tag = classifier.classify(dialogue_act_features(dcell[14:]))
                worksheet.cell(row_idx + 1, 7, tag)
    workbook.save(filename)

print('\nComplete. See "Unclassified" folder for newly classified CSVs.')

end_time = timeit.default_timer()

print('\nRun time: ', str(round((end_time - start_time), 0)), ' seconds')